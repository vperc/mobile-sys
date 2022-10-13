from django.shortcuts import render, redirect,HttpResponse,HttpResponseRedirect
from django.core.validators import RegexValidator
from django.core.validators import ValidationError
from app01 import models
from django import forms
from django.utils.safestring import mark_safe
from app01.utils.pagination import Pagination
from app01.utils.bootstrap import BootStrapModelForm
from app01.utils.form import UserModelForm,PrettyModelForm,PrettyEditModelForm
from django.core.files.uploadhandler import InMemoryUploadedFile
from openpyxl import load_workbook
import os

def depart_list(request):
    """部门列表"""
    # 去数据库中获取所有的部门列表
    # 列表放的一个一个对象
    queryset = models.Department.objects.all()

    page_object = Pagination(request,queryset,page_size=8)

    context = {
        "queryset": page_object.page_queryset,
        "page_string":page_object.html(),
    }

    return render(request, 'depart_list.html', context)

def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户提交过来的数据(title输入为空)
    title = request.POST.get("title")

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回部门表 redirect
    return redirect("/depart/list")

def depart_delete(request):
    """删除部门"""
    # 获取ID
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()

    # 重定向回部门表 redirect
    return redirect("/depart/list")

def depart_edit(request, nid):
    """修改部门"""
    if request.method == "GET":
        # 根据nid,获取他的数据 [obj,]
        row_object = models.Department.objects.filter(id=nid).first()
        # print(row_object.id, row_object.title)
        return render(request, 'depart_edit.html', {'row_object': row_object})
    # 获取用户提交的标题
    title = request.POST.get("title")

    # 根据ID找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)

    # 重定向回部门表 redirect
    return redirect("/depart/list")

def depart_multi(request):
    """ 批量上传 (Excel文件)"""

    #1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    #2.对象传递给openpyxl,由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    #3.循环h获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    #保存到数据库
    if text.is_valid():
        text.save()
        return JsonResponse({"status": True})

    return HttpResponse("上传")
    # if request.method == "POST":
    #     myFile = request.FILES.get('filename',None)
    #     if not myFile:
    #         return HttpResponse('no file for upload!')
    #     f = open(os.path.join('C:\\upload',myFile.name),'wb+')
    #     for chunk in myFile.chunks():
    #         f.write(chunk)
    #     f.close()
    #     return HttpResponse('upload over!')
    # else:
    #     return render(request,'upload_list.html')

import csv
def depart_csv(request):
    "文件的下载"
    # w只写模式，newline=‘’指的是特殊符号不会进行转义
    with open("exp.csv", 'w', newline='') as f_csv:
        writer = csv.writer(f_csv)
        writer.writerows(['1', '2', '3'], )  # 参数为一个列表，调用一次写一行
        writer.writerow(['a', 'b', 'c'])  # 参数为一个列表，调用一次写一行
    print("执行结束")
    # 1.修改响应类型为csv
    response = HttpResponse(content_type='text/csv')
    # 2.在响应里面添加特殊的响应头,添加文件名
    response['Content-Disposition'] = 'attachment;filename="test.csv"'
    # 3.准备数据
    data = list(range(10))
    # 4.调用csv.writer，并直接指向响应对象
    writer = csv.writer(response)
    # 5.写入数据
    writer.writerow(data)
    return response

def depart_up(request):
    if request.method == "GET":
        return render(request, "file_up.html")
    elif request.method == "POST":
        file = request.FILES.get("file")
        title = request.POST.get("title")
        Content.objects.create(title=title,picture=file)
        return HttpResponse("文件上传成功")




